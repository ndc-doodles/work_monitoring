from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.contrib.auth.models import User  
from .models import *
from datetime import datetime
from django.views.decorators.csrf import csrf_exempt
from datetime import date
import openpyxl
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment
from django.contrib.auth.decorators import login_required
from django.core.paginator import Paginator
from django.contrib.auth.hashers import check_password, make_password
import random
from django.core.mail import send_mail
from datetime import time, timedelta
from django.utils.dateparse import parse_datetime
from django.utils.timezone import localtime, make_aware, is_naive
from django.views.decorators.cache import never_cache
from django.utils.cache import add_never_cache_headers

from django.http import JsonResponse












def index(request):
    return render(request,'index.html')


@never_cache
def admin_login(request):
    if request.user.is_authenticated and request.user.is_superuser:
        return redirect('admin_dashboard')

    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)
        if user is not None and user.is_superuser:
            login(request, user)
            return redirect('admin_dashboard')
        else:
            messages.error(request, "Invalid credentials or not a superuser")

    return render(request, 'admin_login.html')

def admin_logout(request):
    request.session.flush()
    return redirect("index")   # 👈 Redirect to your index page, not admin_login

@never_cache
@login_required(login_url='admin_login')
def admin_dashboard(request):
    if request.method == "POST":
        if "add_department" in request.POST:
            name = request.POST.get("department_name")
            if name:
                Department.objects.create(name=name)
                return redirect('admin_dashboard')

        elif "add_team" in request.POST:
            name = request.POST.get("team_name")
            if name:
                Team.objects.create(name=name)
                return redirect('admin_dashboard')

    departments = Department.objects.all()
    teams = Team.objects.all()
    return render(request, "admin_dashboard.html", {"departments": departments, "teams": teams})





def delete_department(request, pk):
    dept = get_object_or_404(Department, pk=pk)
    dept.delete()
    return redirect('admin_dashboard')

def delete_team(request, pk):
    team = get_object_or_404(Team, pk=pk)
    team.delete()
    return redirect('admin_dashboard')


@csrf_exempt

def admin_usermanagement(request):
    users = User.objects.all()
    departments = Department.objects.all()
    teams = Team.objects.all()

    if request.method == "POST":
        user_id = request.POST.get("id")  # for edit, will be None for new users
        name = request.POST.get("name")
        employee_id = request.POST.get("employee_id")
        email = request.POST.get("email")
        phone = request.POST.get("phone")
        department_id = request.POST.get("department")
        team_id = request.POST.get("team")
        job_Position = request.POST.get("job_Position")
        designation = request.POST.get("designation")
        work_location = request.POST.get("work_location")
        username = request.POST.get("username")
        password = request.POST.get("password")
        status = request.POST.get("status")
        image = request.FILES.get("profile_image")
        joining_date = request.POST.get("joining_date")

        department = Department.objects.get(id=department_id) if department_id else None
        team = Team.objects.get(id=team_id) if team_id else None

        if user_id:  # Edit existing user
            user = get_object_or_404(User, id=user_id)

            # Uniqueness checks excluding current user
            if User.objects.filter(employee_id=employee_id).exclude(id=user.id).exists():
                messages.error(request, "⚠️ Employee ID already exists for another user.")
                return redirect("admin_usermanagement")
            if User.objects.filter(username=username).exclude(id=user.id).exists():
                messages.error(request, "⚠️ Username already exists for another user.")
                return redirect("admin_usermanagement")

            # Update user fields
            user.name = name
            user.employee_id = employee_id
            user.email = email
            user.phone = phone
            user.department = department
            user.team = team
            user.job_Position = job_Position
            user.designation = designation
            user.work_location = work_location
            user.username = username
            user.status = status
            if password:
                user.password = make_password(password)
            if image:
                user.profile_image = image
            if joining_date:
                try:
                    user.joining_date = datetime.strptime(joining_date, "%Y-%m-%d").date()
                except ValueError:
                    messages.error(request, "Invalid date format! Use YYYY-MM-DD.")
                    return redirect("admin_usermanagement")
            user.save()
            messages.success(request, "✅ User updated successfully!")
        
        else:  # Create new user
            if User.objects.filter(employee_id=employee_id).exists():
                messages.error(request, "⚠️ Employee ID already exists. Please choose another one.")
                return redirect("admin_usermanagement")
            if User.objects.filter(username=username).exists():
                messages.error(request, "⚠️ Username already exists. Please choose another one.")
                return redirect("admin_usermanagement")

            User.objects.create(
                name=name,
                employee_id=employee_id,
                email=email,
                phone=phone,
                department=department,
                team=team,
                job_Position=job_Position,
                designation=designation,
                work_location=work_location,
                username=username,
                password=make_password(password) if password else "",
                status=status,
                profile_image=image,
                joining_date=datetime.strptime(joining_date, "%Y-%m-%d").date() if joining_date else None
            )
            messages.success(request, "✅ User created successfully!")

        return redirect("admin_usermanagement")

    context = {
        "users": users,
        "departments": departments,
        "teams": teams,
    }
    return render(request, "admin_usermanagement.html", context)


def edit_user(request):
    if request.method == "POST":
        user_id = request.POST.get("id")
        user = get_object_or_404(User, id=user_id)

        user.employee_id = request.POST.get("edit_emp_id")
        user.name = request.POST.get("edit_name")
        user.email = request.POST.get("edit_email")
        user.phone = request.POST.get("edit_phone")

        # Department + Team (FKs via dropdown IDs)
        dept_id = request.POST.get("edit_department")
        team_id = request.POST.get("edit_team")
        user.department = Department.objects.get(id=dept_id) if dept_id else None
        user.team = Team.objects.get(id=team_id) if team_id else None

        user.job_Position = request.POST.get("edit_job_position")
        user.designation = request.POST.get("edit_designation")
        user.work_location = request.POST.get("edit_work_location")
        user.username = request.POST.get("edit_username")
        user.status = request.POST.get("edit_status")

        # Password (hash it)
        password = request.POST.get("edit_password")
        if password:
            user.password = make_password(password)

        # Joining Date
        joining_date = request.POST.get("edit_joining_date")
        if joining_date:
            try:
                user.joining_date = datetime.strptime(joining_date, "%Y-%m-%d").date()
            except ValueError:
                messages.error(request, "Invalid date format! Use YYYY-MM-DD.")
                return redirect("admin_usermanagement")
        else:
            user.joining_date = None

        # Profile Image
        if request.FILES.get("edit_profile_upload"):
            user.profile_image = request.FILES["edit_profile_upload"]

        user.save()
        messages.success(request, "✅ User updated successfully!")
        return redirect("admin_usermanagement")

    return redirect("admin_usermanagement")



def delete_user(request, id):
    user = get_object_or_404(User, id=id)
    user.delete()
    return redirect('admin_usermanagement') 



def admin_chat(request):
    users = User.objects.exclude(id=request.user.id)  # company users
    extra_contacts = ExtraContact.objects.all()       # manually added numbers
    return render(request, "admin_chat.html", {
        "users": users,
        "extra_contacts": extra_contacts
    })

def add_contact(request):
    if request.method == "POST":
        name = request.POST.get("name")
        phone = request.POST.get("phone")

        # prevent duplicate phone numbers
        if ExtraContact.objects.filter(phone=phone).exists():
            messages.error(request, "Contact with this phone already exists!")
        else:
            ExtraContact.objects.create(name=name, phone=phone)
            messages.success(request, "Contact added successfully!")

        return redirect("admin_chat")  # back to chat page

    return redirect("admin_chat")




@never_cache
def login_view(request):
    # If user is already logged in, redirect to dashboard
    if request.session.get("user_id") and request.session.get("position"):
        if request.session["position"] == "team_lead":
            return redirect("teamlead_dashboard")
        else:
            return redirect("teammember_dashboard")

    if request.method == "POST":
        username = request.POST.get("username")
        password = request.POST.get("password")
        position = request.POST.get("position")

        try:
            user = User.objects.get(username=username)
            if user.password == password:
                db_position = user.job_Position.lower().replace(" ", "_")
                if db_position == position:
                    # Mark user as active
                    user.status = "active"
                    user.last_login_time = timezone.now()
                    user.save()

                    # Save session
                    request.session["user_id"] = user.id
                    request.session["position"] = db_position
                    request.session["login_time"] = str(timezone.now())

                    # ✅ Redirect to dashboard to immediately show username in navbar
                    if db_position == "team_lead":
                        return redirect("teamlead_dashboard")
                    else:
                        return redirect("teammember_dashboard")

        except User.DoesNotExist:
            messages.error(request, "Invalid username or password")

    return render(request, "user_login.html")


   
def get_logged_in_user_api(request):
    if request.session.get("user_id"):
        try:
            user = User.objects.get(id=request.session["user_id"])
            return JsonResponse({"name": user.name, "job_position": user.job_Position})
        except User.DoesNotExist:
            pass
    return JsonResponse({"name": None})





def forgot_password(request):
    if request.method == "POST":
        email = request.POST.get("email")

        try:
            user = User.objects.get(email=email)
            otp = random.randint(100000, 999999)
            request.session['reset_email'] = email
            request.session['reset_otp'] = str(otp)

            # Send OTP via email
            send_mail(
                "Password Reset OTP",
                f"Your OTP for password reset is {otp}",
                settings.DEFAULT_FROM_EMAIL,
                [email],
                fail_silently=False,
            )

            messages.success(request, "OTP sent to your email")
            return redirect("verify_otp")

        except User.DoesNotExist:
            messages.error(request, "Email not registered")
    
    return render(request, "forgot_password.html")


# Step 2: Verify OTP
def verify_otp(request):
    if request.method == "POST":
        otp = request.POST.get("otp")
        if otp == request.session.get("reset_otp"):
            return redirect("reset_password")
        else:
            messages.error(request, "Invalid OTP")

    return render(request, "verify_otp.html")


# Step 3: Reset Password
def reset_password(request):
    if request.method == "POST":
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if password == confirm_password:
            email = request.session.get("reset_email")
            user = User.objects.get(email=email)
            user.password = password   # ⚠️ Better: use user.set_password(password) if using Django's auth User
            user.save()

            # Clear session values
            request.session.pop("reset_email", None)
            request.session.pop("reset_otp", None)

            messages.success(request, "Password reset successful. Please login.")
            return redirect("login")
        else:
            messages.error(request, "Passwords do not match")

    return render(request, "reset_password.html")







@never_cache
def teamlead_dashboard(request):
    # Redirect to index if not logged in
    if not request.session.get("user_id") or request.session.get("position") != "team_lead":
        return redirect("index")

    try:
        team_lead = User.objects.get(id=request.session['user_id'])
    except User.DoesNotExist:
        request.session.flush()
        return redirect("index")

    # Login time
    login_time_str = request.session.get('login_time')
    login_time = parse_datetime(login_time_str) if login_time_str else None
    if login_time:
        if is_naive(login_time):
            login_time = make_aware(login_time)
        login_time = localtime(login_time)

    # Post new announcement
    if request.method == "POST":
        message = request.POST.get("message")
        if message and message.strip():
            Announcement.objects.create(
                title="Announcement",
                message=message,
                created_by=team_lead,
                created_at=timezone.now()
            )
        return redirect('teamlead_dashboard')

    # Team members data
    team_members = User.objects.filter(team=team_lead.team)
    total_users = team_members.count()
    active_members = team_members.filter(status="active").count()
    inactive_members = team_members.filter(status="inactive").count()

    # Most recent login
    last_login_user = team_members.filter(last_login_time__isnull=False).order_by('-last_login_time').first()
    last_login = localtime(last_login_user.last_login_time) if last_login_user else None

    # Announcements from last 12 hours
    cutoff = timezone.now() - timedelta(hours=12)
    announcements = Announcement.objects.filter(
        created_by=team_lead,
        created_at__gte=cutoff
    ).order_by('-created_at')

    # Format login/logout times
    for member in team_members:
        if member.last_login_time:
            member.last_login_time = localtime(member.last_login_time)
        if member.last_logout_time:
            member.last_logout_time = localtime(member.last_logout_time)

    response = render(request, 'teamlead_dashboard.html', {
        'total_users': total_users,
        'active_members': active_members,
        'inactive_members': inactive_members,
        'team_members': team_members,
        'announcements': announcements,
        'login_time': login_time,
        'last_login': last_login,
    })

    # 🔒 Block browser from caching this page
    add_never_cache_headers(response)
    return response

def is_within_time_range(start_time, end_time, now=None):
    now = now or timezone.localtime().time()
    return start_time <= now <= end_time







def teamlead_chat(request):
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("login_view")

    current_user = User.objects.get(id=user_id)

    # All users
    users = list(User.objects.all())

    # Reorder: put current_user first
    users.sort(key=lambda u: 0 if u.id == current_user.id else 1)

    extra_contacts = ExtraContact.objects.all()

    return render(request, 'teamlead_chat.html', {
        'current_user': current_user,
        'users': users,
        'extra_contacts': extra_contacts,
    })





def teammember_chat(request):
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("login_view")  

    current_user = User.objects.get(id=user_id)

    users = list(User.objects.exclude(id=user_id))  # Other users
    users.insert(0, current_user)  # Current user on top

    extra_contacts = ExtraContact.objects.all()

    return render(request, 'teammember_chat.html', {
        'current_user': current_user,
        'users': users,
        'extra_contacts': extra_contacts,
        'role': 'teammember',
    })






@never_cache
def teammember_dashboard(request):
    # ❌ Redirect if not logged in
    if not request.session.get("user_id") or request.session.get("position") != "team_member":
        return redirect("login")

    try:
        team_member = User.objects.get(id=request.session['user_id'])
    except User.DoesNotExist:
        request.session.flush()
        return redirect("login")

    team_name = team_member.team

    # Login time
    login_time_str = request.session.get('login_time')
    login_time = parse_datetime(login_time_str) if login_time_str else None
    if login_time:
        if is_naive(login_time):
            login_time = make_aware(login_time)
        login_time = localtime(login_time)

    # Announcements from last 12 hours
    cutoff = now() - timedelta(hours=12)
    announcements = Announcement.objects.filter(
        created_by__team=team_name,
        created_at__gte=cutoff
    ).order_by('-created_at')

    # Morning/Evening time windows
    morning_start, morning_end = time(9, 30), time(10, 30)
    evening_start, evening_end = time(17, 0), time(18, 15)

    if request.method == "POST":
        now_time = localtime().time()

        # Morning report
        if 'morning_submit' in request.POST:
            if is_within_time_range(morning_start, morning_end, now_time):
                report_text = request.POST.get("morning_report")
                status = request.POST.get("morning_status")
                if report_text and status:
                    MorningReport.objects.create(
                        user=team_member,
                        department=team_member.department.name if team_member.department else None,
                        team=team_member.team.name if team_member.team else None,
                        report_text=report_text,
                        status=status
                    )
                    messages.success(request, "Morning report submitted successfully.")
                    return redirect('teammember_dashboard')
            else:
                messages.error(request, "You can only submit morning reports between 9:30 and 10:30 AM.")

        # Evening report
        elif 'evening_submit' in request.POST:
            if is_within_time_range(evening_start, evening_end, now_time):
                report_text = request.POST.get("evening_report")
                status = request.POST.get("evening_status")
                if report_text and status:
                    EveningReport.objects.create(
                        user=team_member,
                        department=team_member.department.name if team_member.department else None,
                        team=team_member.team.name if team_member.team else None,
                        report_text=report_text,
                        status=status
                    )
                    messages.success(request, "Evening report submitted successfully.")
                    return redirect('teammember_dashboard')
            else:
                messages.error(request, "You can only submit evening reports between 5:00 and 6:15 PM.")

    # ✅ Fetch reports submitted in last 24 hours
    report_cutoff = now() - timedelta(hours=24)

    morning_reports = MorningReport.objects.filter(
        user=team_member,
        created_at__gte=report_cutoff
    ).values("report_text", "status", "created_at")
    for r in morning_reports:
        r["type"] = "Morning"

    evening_reports = EveningReport.objects.filter(
        user=team_member,
        created_at__gte=report_cutoff
    ).values("report_text", "status", "created_at")
    for r in evening_reports:
        r["type"] = "Evening"

    all_reports = list(morning_reports) + list(evening_reports)
    all_reports = sorted(all_reports, key=lambda x: x["created_at"], reverse=True)

    return render(request, 'teammember_dashboard.html', {
        'announcements': announcements,
        'morning_allowed': is_within_time_range(morning_start, morning_end),
        'evening_allowed': is_within_time_range(evening_start, evening_end),
        'login_time': login_time,
        'all_reports': all_reports,  # ✅ send to template
    })

@never_cache
def teamlead_reports(request):
    # ✅ Prevent KeyError if user is logged out
    if not request.session.get("user_id") or request.session.get("position") != "team_lead":
        return redirect("index")

    team_lead = get_object_or_404(User, id=request.session["user_id"])
    team_name = team_lead.team
    show_all = request.GET.get("all") == "1"

    if show_all:
        morning_reports = MorningReport.objects.filter(team=team_name).order_by("-created_at")
        evening_reports = EveningReport.objects.filter(team=team_name).order_by("-created_at")
    else:
        today = date.today()
        morning_reports = MorningReport.objects.filter(team=team_name, created_at__date=today)
        evening_reports = EveningReport.objects.filter(team=team_name, created_at__date=today)

    if "export" in request.GET:
        if not morning_reports.exists() and not evening_reports.exists():
            messages.error(request, "No reports available to export.")
            return redirect("teamlead_reports")
        else:
            wb = Workbook()
            ws = wb.active
            ws.title = "Team Reports"

            headers = ["Date", "Time", "Report Type", "Name", "Team", "Department", "Report", "Status"]
            ws.append(headers)

            def add_reports(report_list, report_type):
                for report in report_list:
                    local_time = timezone.localtime(report.created_at)
                    ws.append([
                        local_time.strftime("%Y-%m-%d"),
                        local_time.strftime("%I:%M %p"),
                        report_type,
                        getattr(report.user, "name", report.user.username),
                        report.team,
                        report.department,
                        report.report_text,
                        report.status
                    ])

            add_reports(morning_reports, "Morning")
            add_reports(evening_reports, "Evening")

            for row in ws.iter_rows(min_row=2, min_col=7, max_col=7):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            column_widths = [15, 12, 12, 20, 15, 20, 50, 15]
            for i, width in enumerate(column_widths, start=1):
                ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = width

            response = HttpResponse(
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = 'attachment; filename="team_reports.xlsx"'
            wb.save(response)
            return response

    return render(request, "teamlead_reports.html", {
        "morning_reports": morning_reports,
        "evening_reports": evening_reports,
        "show_all": show_all,
    })
@never_cache
def teamlead_project_assigning(request):
    # 🔒 Prevent KeyError: check session before accessing user
    if not request.session.get("user_id") or request.session.get("position") != "team_lead":
        return redirect("index")

    team_lead = get_object_or_404(User, id=request.session["user_id"])
    departments = Department.objects.all()

    team_members = User.objects.filter(
        team=team_lead.team,
        job_Position__icontains="team member"
    ).exclude(id=team_lead.id)

    if request.method == "POST":
        # First create the project without files/images
        project = ProjectAssign.objects.create(
            team=team_lead.team,
            department_id=request.POST.get("department"),
            assign_to_id=request.POST.get("assign_to"),
            assigned_by=team_lead,
            work_name=request.POST.get("work_name"),
            work_type=request.POST.get("work_type"),
            category=request.POST.get("category"),
            description=request.POST.get("description"),
            deadline=request.POST.get("deadline"),
            additional_notes=request.POST.get("additional_notes"),
            color_preference=request.POST.get("color_preference"),
            content_example=request.POST.get("content_example"),
            priority=request.POST.get("priority"),
        )

        # Handle multiple uploaded files
        for file in request.FILES.getlist("upload_file[]"):
            ProjectFile.objects.create(project=project, file=file)

        # Handle multiple uploaded images
        for image in request.FILES.getlist("upload_image[]"):
            ProjectImage.objects.create(project=project, image=image)

        return redirect("teamlead_project_assigning")

    projects = ProjectAssign.objects.filter(team=team_lead.team)

    return render(request, "teamlead_project_assigning.html", {
        "departments": departments,
        "team_lead": team_lead,
        "team_members": team_members,
        "projects": projects,
    })

def project_assign_edit(request, pk):
    project = get_object_or_404(ProjectAssign, id=pk)
    departments = Department.objects.all()
    team_members = User.objects.filter(
        team=project.team, job_Position__iexact="Team Member"
    ).exclude(id=project.assigned_by.id)

    if request.method == "POST":
        project.department_id = request.POST.get("department")
        project.assign_to_id = request.POST.get("assign_to")
        project.work_name = request.POST.get("work_name")
        project.work_type = request.POST.get("work_type")
        project.category = request.POST.get("category")
        project.description = request.POST.get("description")

        deadline = request.POST.get("deadline")
        project.deadline = deadline if deadline else None

        project.additional_notes = request.POST.get("additional_notes")
        project.color_preference = request.POST.get("color_preference")
        project.content_example = request.POST.get("content_example")
        project.priority = request.POST.get("priority")
        project.save()

        # ✅ Save new files (without removing old ones)
        for file in request.FILES.getlist("upload_file[]"):
            ProjectFile.objects.create(project=project, file=file)

        # ✅ Save new images (without removing old ones)
        for image in request.FILES.getlist("upload_image[]"):
            ProjectImage.objects.create(project=project, image=image)

        return redirect("teamlead_project_assigning")

    return render(request, "teamlead_project_assigning.html", {
        "project": project,
        "departments": departments,
        "team_members": team_members,
    })

def project_assign_delete(request, pk):
    project = get_object_or_404(ProjectAssign, id=pk)
    project.delete()
    return redirect("teamlead_project_assigning")

@never_cache
def teammember_project(request):
    user_id = request.session.get("user_id")
    if not user_id:   # if session expired or user logged out
        return redirect("index")

    user = get_object_or_404(User, id=user_id)
    projects = ProjectAssign.objects.filter(assign_to=user).order_by("-assigned_date")

    if request.method == "POST":
        project_id = request.POST.get("project_id")
        status = request.POST.get("status")
        project = get_object_or_404(ProjectAssign, id=project_id, assign_to=user)
        project.status = status
        project.save()
        return redirect("teammember_project")

    return render(request, "teammember_project.html", {"projects": projects})
def update_project_status(request, pk):
    user_id = request.session.get("user_id")  
    if not user_id:
        messages.error(request, "You must be logged in to update status.")
        return redirect("index")  # or your login page

    user = get_object_or_404(User, id=user_id)
    project = get_object_or_404(ProjectAssign, pk=pk, assign_to=user)

    if request.method == "POST":
        new_status = request.POST.get("status")
        if new_status in dict(ProjectAssign.STATUS_CHOICES):
            project.status = new_status
            project.save()
            messages.success(request, f"Project status updated to {new_status}")
        else:
            messages.error(request, "Invalid status selected.")

    return redirect("teammember_project")
@never_cache
def teamlead_notepad(request):
    # 🔒 Prevent access if logged out
    user_id = request.session.get("user_id")
    if not user_id or request.session.get("position") != "team_lead":
        return redirect("index")

    user = get_object_or_404(User, id=user_id)

    # ✅ Only this user's notes
    all_notes = Notepad.objects.filter(user=user).order_by("-updated_at")

    # Pagination
    paginator = Paginator(all_notes, 4)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # View/Edit a specific note (only if it belongs to this user)
    note = None
    note_id = request.GET.get("note_id")
    if note_id:
        note = Notepad.objects.filter(id=note_id, user=user).first()

    if request.method == "POST":
        note_id_post = request.POST.get("note_id")
        title = request.POST.get("title") or "Untitled"
        content = request.POST.get("content")

        if note_id_post:
            # ✅ Update only if owned by this user
            note = get_object_or_404(Notepad, id=note_id_post, user=user)
            note.title = title
            note.content = content
            note.save()
        else:
            # ✅ New note tied to this user
            note = Notepad.objects.create(user=user, title=title, content=content)

        return redirect(f"{request.path}?note_id={note.id}")

    return render(
        request,
        "teamlead_notepad.html",
        {"note": note, "page_obj": page_obj},
    )




@never_cache
def teammember_notepad(request):
    user_id = request.session.get("user_id")
    if not user_id:   # 🔒 No session → back to index
        return redirect("index")

    user = get_object_or_404(User, id=user_id)

    # All notes for this user
    all_notes = Notepad.objects.filter(user=user).order_by("-updated_at")

    # Pagination (4 notes per page)
    paginator = Paginator(all_notes, 4)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    # Note to edit/view
    note_id = request.GET.get("note_id")
    note = None
    if note_id:
        try:
            note = Notepad.objects.get(id=note_id, user=user)
        except Notepad.DoesNotExist:
            note = None  # Invalid ID → open new note

    if request.method == "POST":
        note_id_post = request.POST.get("note_id")
        title = request.POST.get("title") or "Untitled"
        content = request.POST.get("content")

        if note_id_post:
            # Update existing note
            note = get_object_or_404(Notepad, id=note_id_post, user=user)
            note.title = title
            note.content = content
            note.save()
        else:
            # Create new note
            note = Notepad.objects.create(user=user, title=title, content=content)

        return redirect(f"{request.path}?note_id={note.id}")

    return render(request, "teammember_notepad.html", {"note": note, "page_obj": page_obj})




@never_cache
def teamlead_repository(request):
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("index")  # or "login_view" if you want login page

    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        title = request.POST.get("title")
        description = request.POST.get("description")
        link = request.POST.get("link")
        file = request.FILES.get("file")

        Knowledge.objects.create(
            department=user.department,
            user=user,
            title=title,
            description=description,
            link=link,
            file=file
        )
        return redirect("teamlead_repository")

    knowledge_items = Knowledge.objects.filter(department=user.department).order_by("-created_at")

    response = render(request, "teamlead_repository.html", {"knowledge_items": knowledge_items})
    
    # Extra cache protection
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    
    return response


def teamlead_repository_delete(request, pk):
    user = get_object_or_404(User, id=request.session.get("user_id"))
    resource = get_object_or_404(Knowledge, id=pk, department=user.department)

    if request.method == "POST":
        resource.delete()
        return redirect("teamlead_repository")

from django.shortcuts import redirect, get_object_or_404
from django.views.decorators.cache import never_cache

@never_cache
def teammember_repository(request):
    user_id = request.session.get("user_id")
    if not user_id:   # 🔒 No session → redirect to index
        return redirect("index")

    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        title = request.POST.get("title")
        description = request.POST.get("description")
        link = request.POST.get("link")
        file = request.FILES.get("file")

        Knowledge.objects.create(
            department=user.department,
            user=user,
            title=title,
            description=description,
            link=link,
            file=file
        )
        return redirect("teammember_repository")

    knowledge_items = Knowledge.objects.filter(department=user.department).order_by("-created_at")

    return render(request, "teammember_repository.html", {"knowledge_items": knowledge_items})



@login_required
def teammember_repository_delete(request, pk):
    user = get_object_or_404(User, id=request.session.get("user_id"))
    resource = get_object_or_404(Knowledge, id=pk, department=user.department)

    if request.method == "POST":
        resource.delete()
        return redirect("teammember_repository")



@never_cache
def teamlead_profile(request):
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("index")  # User not logged in, go to index page

    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        action = request.POST.get("action")

        # Change Password
        if action == "change_password":
            current_password = request.POST.get("current_password")
            new_password = request.POST.get("new_password")
            confirm_password = request.POST.get("confirm_password")

            if not check_password(current_password, user.password):
                messages.error(request, "Current password is incorrect.")
            elif new_password != confirm_password:
                messages.error(request, "New password and confirmation do not match.")
            else:
                user.password = make_password(new_password)
                user.save()
                messages.success(request, "Password changed successfully!")

        # Edit Profile
        elif action == "edit_profile":
            user.name = request.POST.get("name")
            user.email = request.POST.get("email")
            user.phone = request.POST.get("phone")
            user.work_location = request.POST.get("work_location")

            if "profile_image" in request.FILES:
                user.profile_image = request.FILES["profile_image"]

            user.save()
            messages.success(request, "Profile updated successfully!")

        return redirect("teamlead_profile")

    # Render page with cache prevention
    response = render(request, "teamlead_profile.html", {"user": user})
    response["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


@never_cache
def teammember_profile(request):
    user_id = request.session.get("user_id")
    if not user_id:   # 🔒 Not logged in → back to index
        return redirect("index")

    user = get_object_or_404(User, id=user_id)

    if request.method == "POST":
        action = request.POST.get("action")

        # ✅ Change Password
        if action == "change_password":
            current_password = request.POST.get("current_password")
            new_password = request.POST.get("new_password")
            confirm_password = request.POST.get("confirm_password")

            if not check_password(current_password, user.password):
                messages.error(request, "Current password is incorrect.")
            elif new_password != confirm_password:
                messages.error(request, "New password and confirmation do not match.")
            else:
                user.password = make_password(new_password)
                user.save()
                messages.success(request, "Password changed successfully!")

        # ✅ Edit Profile
        elif action == "edit_profile":
            user.name = request.POST.get("name")
            user.email = request.POST.get("email")
            user.phone = request.POST.get("phone")
            user.work_location = request.POST.get("work_location")

            if "profile_image" in request.FILES:
                user.profile_image = request.FILES["profile_image"]

            user.save()
            messages.success(request, "Profile updated successfully!")

        return redirect("teammember_profile")

    return render(request, "teammember_profile.html", {"user": user})

@never_cache
def teammember_task(request):
    user_id = request.session.get("user_id")
    if not user_id:   # 🔒 session expired or logged out
        return redirect("index")   # safer than showing error

    user = get_object_or_404(User, id=user_id)

    # ✅ Only tasks created by this logged-in user
    tasks = Task.objects.filter(created_by=user).order_by("-created_at")

    if request.method == "POST":
        title = request.POST.get("title")
        description = request.POST.get("description")
        if title:  # prevent empty titles
            Task.objects.create(
                title=title,
                description=description,
                assigned_to=user,   # optional: assign to self
                created_by=user     # track who created it
            )
        return redirect("teammember_task")

    return render(request, "teammember_task.html", {"tasks": tasks})


# TEAM MEMBER UPDATE TASK
def update_task(request, task_id):
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("login_view")

    user = User.objects.get(id=user_id)

    # ✅ Only allow updating tasks created by this user
    task = get_object_or_404(Task, id=task_id, created_by=user)

    if request.method == "POST":
        status = request.POST.get("status")
        task.status = status
        # Update progress based on status
        task.progress = {"pending": 0, "in_progress": 50, "completed": 100}.get(status, 0)
        task.save()

    return redirect("teammember_task")


# TEAM MEMBER DELETE TASK
def delete_task(request, task_id):
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("login_view")

    user = User.objects.get(id=user_id)

    # ✅ Only allow deletion of tasks created by this user
    task = get_object_or_404(Task, id=task_id, created_by=user)

    if request.method == "POST":
        task.delete()

    return redirect("teammember_task")



@never_cache
def teamlead_task(request):
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("index")  # Redirect to index if not logged in

    user = get_object_or_404(User, id=user_id)

    # Only show tasks created by this logged-in user
    tasks = Task.objects.filter(created_by=user).order_by('-created_at')

    if request.method == "POST":
        title = request.POST.get("title")
        description = request.POST.get("description")
        if title:  # prevent creating empty tasks
            Task.objects.create(
                title=title,
                description=description,
                assigned_to=user,
                created_by=user
            )
        return redirect("teamlead_task")

    response = render(request, 'teamlead_task.html', {"tasks": tasks})
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
    response['Pragma'] = 'no-cache'
    response['Expires'] = '0'
    return response


def update_task_teamlead(request, task_id):
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("login_view")

    user = User.objects.get(id=user_id)
    # ✅ Only allow updating tasks created by logged-in user
    task = get_object_or_404(Task, id=task_id, created_by=user)

    if request.method == "POST":
        status = request.POST.get("status")
        task.status = status
        task.progress = {"pending": 0, "in_progress": 50, "completed": 100}.get(status, 0)
        task.save()

    return redirect('teamlead_task')


def delete_task_teamlead(request, task_id):
    user_id = request.session.get("user_id")
    if not user_id:
        return redirect("login_view")

    user = User.objects.get(id=user_id)
    # ✅ Only allow deletion of tasks created by logged-in user
    task = get_object_or_404(Task, id=task_id, created_by=user)

    if request.method == "POST":
        task.delete()

    return redirect("teamlead_task")

@never_cache
def teamlead_logout(request):
    if "user_id" in request.session:
        try:
            user = User.objects.get(id=request.session["user_id"])
            user.status = "inactive"
            user.last_logout_time = timezone.now()
            user.save()
        except User.DoesNotExist:
            pass

    request.session.flush()
    response = redirect("index")
    response["Cache-Control"] = "no-cache, no-store, must-revalidate"
    response["Pragma"] = "no-cache"
    response["Expires"] = "0"
    return response


def teammember_logout(request):
    if "user_id" in request.session:
        try:
            user = User.objects.get(id=request.session["user_id"])
            user.status = "inactive"
            user.last_logout_time = timezone.now()
            user.save()
        except User.DoesNotExist:
            pass

    # ✅ Clear session fully
    request.session.flush()

    # ✅ Redirect to index (not login page)
    return redirect("index")